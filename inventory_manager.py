from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from listing_utils import get_listing_variants


HEADERS = [
    "SKU",
    "Card Name",
    "Set",
    "Number",
    "Finish",
    "Rarity",
    "Quantity",
    "Price",
    "Image",
]


def update_inventory(cards, set_id):

    inventory_dir = Path("inventory")
    inventory_dir.mkdir(exist_ok=True)

    inventory_file = inventory_dir / "OPS_Inventory.xlsx"

    # -----------------------------
    # Create or load workbook
    # -----------------------------
    if inventory_file.exists():

        wb = load_workbook(inventory_file)
        ws = wb.active

    else:

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"

        ws.append(HEADERS)

        fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78",
        )

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill

    # -----------------------------
    # Existing SKUs
    # -----------------------------
    existing = {
        row[0]
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[0]
    }

    added = 0
    skipped = 0

    # -----------------------------
    # Import listings
    # -----------------------------
    for card in cards:

        listing = get_listing_variants(card, set_id)

        for variant in listing["variants"]:

            sku = variant["sku"]

            if sku in existing:
                skipped += 1
                continue

            price = variant.get("price", 0)
            

            if price is None:
                price = 0

            try:
                price = float(price)
            except (TypeError, ValueError):
                price = 0

            ws.append([
                sku,
                listing["name"],
                listing["set_name"],
                listing["number"],
                variant["finish"],
                listing["rarity"],
                1,
                price,
                variant["image"],
            ])

            existing.add(sku)
            added += 1

    # -----------------------------
    # Format Price column
    # -----------------------------
    for cell in ws["H"][1:]:
        cell.number_format = "$0.00"

    # -----------------------------
    # Auto-size columns
    # -----------------------------
    for column in ws.columns:

        width = 0
        letter = column[0].column_letter

        for cell in column:

            if cell.value is None:
                continue

            width = max(width, len(str(cell.value)))

        ws.column_dimensions[letter].width = min(width + 2, 40)

    ws.freeze_panes = "A2"

    wb.save(inventory_file)

    print("\n===================================")
    print(" Inventory Updated")
    print("===================================")
    print(f"Added      : {added}")
    print(f"Skipped    : {skipped}")
    print(f"Total SKUs : {len(existing)}")