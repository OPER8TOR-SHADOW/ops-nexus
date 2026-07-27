from pathlib import Path

from openpyxl import load_workbook


def load_inventory_rows(path=None):
    inventory_path = Path(path) if path else Path("inventory") / "OPS_Inventory.xlsx"

    if not inventory_path.exists():
        return []

    workbook = load_workbook(inventory_path, data_only=True)
    sheet = workbook.active

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        rows.append(
            {
                "sku": str(row[0]).strip() if row[0] is not None else "",
                "card_name": str(row[1]).strip() if len(row) > 1 and row[1] is not None else "",
                "set_name": str(row[2]).strip() if len(row) > 2 and row[2] is not None else "",
                "number": str(row[3]).strip() if len(row) > 3 and row[3] is not None else "",
                "finish": str(row[4]).strip() if len(row) > 4 and row[4] is not None else "",
                "rarity": str(row[5]).strip() if len(row) > 5 and row[5] is not None else "",
                "quantity": int(row[6]) if len(row) > 6 and row[6] is not None else 0,
                "price": float(row[7]) if len(row) > 7 and row[7] is not None else 0.0,
                "image": str(row[8]).strip() if len(row) > 8 and row[8] is not None else "",
            }
        )

    return rows
