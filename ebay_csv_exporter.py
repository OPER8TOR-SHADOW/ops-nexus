from pathlib import Path
import csv
from openpyxl import load_workbook


def generate_csv():

    inventory = Path("inventory") / "OPS_Inventory.xlsx"

    if not inventory.exists():
        print("Inventory not found.")
        return

    wb = load_workbook(inventory, data_only=True)
    ws = wb.active

    output = Path("output")
    output.mkdir(exist_ok=True)

    csv_file = output / "ebay_upload.csv"

    with open(csv_file, "w", newline="", encoding="utf-8-sig") as file:

        writer = csv.writer(file)

        writer.writerow([
            "Custom Label (SKU)",
            "Title",
            "Price",
            "Quantity",
            "Condition"
        ])

        for row in ws.iter_rows(min_row=2, values_only=True):

            sku = row[0]
            name = row[1]
            finish = row[4]
            rarity = row[5]
            qty = row[6]
            price = row[7]

            title = f"{name} {finish}"

            writer.writerow([
                sku,
                title,
                price,
                qty,
                "3000"
            ])

    print(f"\n✅ CSV created:\n{csv_file}")