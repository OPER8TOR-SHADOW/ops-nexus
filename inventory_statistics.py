from pathlib import Path

from openpyxl import load_workbook


def show_statistics():

    inventory_file = Path("inventory") / "OPS_Inventory.xlsx"

    if not inventory_file.exists():
        print("\n❌ Inventory not found.")
        input("\nPress Enter to continue...")
        return

    wb = load_workbook(inventory_file, data_only=True)
    ws = wb.active

    total_listings = 0
    inventory_value = 0.0

    finish_counts = {}
    rarity_counts = {}

    for row in ws.iter_rows(min_row=2, values_only=True):

        if not row or row[0] is None:
            continue

        total_listings += 1

        finish = row[4] if len(row) > 4 and row[4] else "Unknown"
        rarity = row[5] if len(row) > 5 and row[5] else "Unknown"

        quantity = row[6]
        if quantity is None:
            quantity = 0

        price = row[7]
        if price is None:
            price = 0.0

        try:
            quantity = int(quantity)
        except (TypeError, ValueError):
            quantity = 0

        try:
            price = float(price)
        except (TypeError, ValueError):
            price = 0.0

        finish_counts[finish] = finish_counts.get(finish, 0) + 1
        rarity_counts[rarity] = rarity_counts.get(rarity, 0) + 1

        inventory_value += quantity * price

    print("\n" + "=" * 45)
    print("       INVENTORY STATISTICS")
    print("=" * 45)

    print(f"\nTotal Listings : {total_listings}")
    print(f"Inventory Value: ${inventory_value:.2f}")

    print("\nFinish Breakdown")
    print("-" * 45)

    for finish in sorted(finish_counts):
        print(f"{finish:<20}{finish_counts[finish]}")

    print("\nRarity Breakdown")
    print("-" * 45)

    for rarity in sorted(rarity_counts):
        print(f"{rarity:<30}{rarity_counts[rarity]}")

    input("\nPress Enter to continue...")