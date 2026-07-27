from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import DEFAULT_QUANTITY, TITLE_FORMAT
from listing_utils import get_listing_variants

BASE_DIR = Path(__file__).resolve().parent


def create_output_file(cards):

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Listings"

    headers = [
        "SKU",
        "Title",
        "Finish",
        "Card Name",
        "Set",
        "Card Number",
        "Rarity",
        "Price",
        "Quantity",
    ]

    # -----------------------------
    # Header
    # -----------------------------
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78",
        )
        cell.alignment = Alignment(horizontal="center")

    row = 2

    printed_total = cards[0]["set"]["printedTotal"]
    set_id = cards[0]["set"]["id"].upper()

    # -----------------------------
    # Generate Listings
    # -----------------------------
    for card in cards:

        listing = get_listing_variants(card, set_id)

        title = TITLE_FORMAT.format(
            name=listing["name"],
            number=listing["number"],
            set_size=printed_total,
            rarity=listing["rarity"],
        )

        for variant in listing["variants"]:

            listing_title = title

            if variant["finish"] == "Reverse Holo":
                listing_title += " Reverse Holo"

            values = [
                variant["sku"],
                listing_title,
                variant["finish"],
                listing["name"],
                listing["set_name"],
                listing["number"],
                listing["rarity"],
                variant["price"],
                DEFAULT_QUANTITY,
            ]

            for col, value in enumerate(values, start=1):
                sheet.cell(row=row, column=col).value = value

            row += 1

    # -----------------------------
    # Price formatting
    # -----------------------------
    for cell in sheet["H"][1:]:
        cell.number_format = "$0.00"

    # -----------------------------
    # Auto-size columns
    # -----------------------------
    for column in sheet.columns:

        max_length = 0
        letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[letter].width = min(max_length + 2, 60)

    # -----------------------------
    # Excel Table
    # -----------------------------
    end_row = row - 1

    table = Table(
        displayName="PokemonListings",
        ref=f"A1:I{end_row}",
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    table.tableStyleInfo = style
    sheet.add_table(table)

    sheet.freeze_panes = "A2"

    output_dir = BASE_DIR / "output"
    output_dir.mkdir(exist_ok=True)

    output_file = output_dir / "Pokemon_eBay_Upload.xlsx"

    workbook.save(output_file)

    print(f"\n✅ Created {row-2} listings!")
    print(f"📄 Saved to: {output_file}")