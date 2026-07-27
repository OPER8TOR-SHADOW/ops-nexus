from pathlib import Path
from openpyxl import load_workbook


def search_inventory():
    inventory_file = Path("inventory") / "OPS_Inventory.xlsx"

    if not inventory_file.exists():
        print("\n❌ Inventory not found.")
        input("\nPress Enter to continue...")
        return

    wb = load_workbook(inventory_file, data_only=True)
    ws = wb.active

    search = input("\nSearch card name or SKU: ").strip().lower()

    results = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[0])
        name = str(row[1])
        finish = str(row[4])
        rarity = str(row[5])
        qty = row[6]
        price = row[7]

        if search in sku.lower() or search in name.lower():
            results.append((sku, name, finish, rarity, qty, price))

    print()

    if not results:
        print("No matching cards found.")
    else:
        print("=" * 100)
        print(f"{'SKU':<15}{'Card Name':<35}{'Finish':<18}{'Rarity':<22}{'Qty':<6}{'Price'}")
        print("=" * 100)

        for sku, name, finish, rarity, qty, price in sorted(results):
            print(f"{sku:<15}{name:<35}{finish:<18}{rarity:<22}{qty:<6}${price:.2f}")

        print("=" * 100)
        print(f"{len(results)} listing(s) found.")

    input("\nPress Enter to continue...")