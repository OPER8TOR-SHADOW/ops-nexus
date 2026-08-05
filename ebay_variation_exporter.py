import csv
import sys
from builtins import (
    FileNotFoundError,
    PermissionError,
    TypeError,
    ValueError,
    enumerate,
    float,
    int,
    isinstance,
    len,
    list,
    open,
    print,
    str,
)
from pathlib import Path

from openpyxl import load_workbook

from config import GITHUB_OWNER, GITHUB_REPO, GITHUB_BRANCH
from listing_utils import get_listing_variants
from pokemon_api import get_cards
from sets import get_set_config_by_id

# ==========================================================
# OPS COLLECTABLES
# eBay Variation Exporter v2.0
# ==========================================================

# -----------------------------
# FILES
# -----------------------------

TEMPLATE = "templates/Carduploader csv template.csv"
INVENTORY = "inventory/OPS_Inventory.xlsx"
OUTPUT_DIR = "output"

# -----------------------------
# LISTING SETTINGS
# -----------------------------

DESCRIPTION = "Thanks for viewing my listing!"
CATEGORY = "183454"
LOCATION = "AUS"

CONDITION_ID = "4000"
CARD_CONDITION = "Near Mint or Better: -(ID: 400010)"
CARD_CONDITION_LABEL = "Near Mint or Better:"

PAYMENT_PROFILE = "BUY IT NOW"
RETURN_PROFILE = "30 DAY RETURN BP"
SHIPPING_PROFILE = "FREE POSTAGE"


def emit_csv_progress(processed, total, current_card):
    print(f"[CSV] {processed}/{total}", flush=True)
    print(f"CURRENT={current_card}", flush=True)

PARENT_GAME = "Pokémon TCG"
PARENT_CARD_TYPE = "Pokemon"
PARENT_MANUFACTURER = "The Pokémon Company"
PARENT_GRADED = "No"
PARENT_CARD_SIZE = "Standard"
PARENT_LANGUAGE = "English"
PARENT_COUNTRY_MANUFACTURE = "Australia"
PARENT_COUNTRY_ORIGIN = "Australia"
PARENT_AGE_LEVEL = "6+"
PARENT_MATERIAL = "Card Stock"
PARENT_POSTAL_CODE = "2830"
PARENT_WEIGHT_MAJOR = "0"
PARENT_WEIGHT_MINOR = "1"
PARENT_WEIGHT_UNIT = "kg"

# ==========================================================
# TEMPLATE
# ==========================================================


def normalize_column_name(value):
    if value is None:
        return ""
    return value.replace("*", "").replace(" ", "").replace("﻿", "").strip().lower()


def load_template():
    with open(TEMPLATE, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        raise ValueError(f"Template is empty: {TEMPLATE}")

    header = rows[0]
    columns = {
        normalize_column_name(value): i
        for i, value in enumerate(header)
        if value is not None
    }

    return header, columns


# ==========================================================
# INVENTORY
# ==========================================================


def format_price(price):
    if price is None or price == "":
        return ""
    try:
        price_value = float(price)
    except (TypeError, ValueError):
        return str(price)
    if price_value.is_integer():
        return str(int(price_value))
    return f"{price_value:.2f}".rstrip("0").rstrip(".")


def load_inventory(selected_set):
    workbook = load_workbook(INVENTORY, data_only=True)
    sheet = workbook.active
    cards = []
    selected_prefix = str(selected_set or "").upper()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        sku = str(row[0]).strip() if row[0] else ""
        if selected_prefix and not sku.upper().startswith(f"{selected_prefix}-"):
            continue

        cards.append({
            "sku": sku,
            "name": str(row[1]).strip() if row[1] else "",
            "set": str(row[2]).strip() if row[2] else "",
            "number": str(row[3]).strip() if row[3] else "",
            "finish": str(row[4]).strip() if row[4] else "",
            "rarity": str(row[5]).strip() if row[5] else "",
            "qty": str(int(row[6])) if isinstance(row[6], (int, float)) and row[6] == int(row[6]) else str(row[6]) if row[6] is not None else "",
            "price": row[7],
            "image": str(row[8]).strip() if row[8] else "",
        })

    return cards


def load_cards_from_api(set_context):
    cards = []
    api_cards = get_cards(set_context["api_set"])

    for card in api_cards:
        listing = get_listing_variants(card, set_context["id"])

        for variant in listing["variants"]:
            cards.append({
                "sku": variant["sku"],
                "name": listing["name"],
                "set": listing["set_name"],
                "number": listing["number"],
                "finish": variant["finish"],
                "rarity": listing["rarity"],
                "qty": "1",
                "price": variant["price"],
                "image": variant["image"],
            })

    return cards


def load_cards_for_export(set_context):
    inventory_path = Path(INVENTORY)

    if inventory_path.exists():
        inventory_cards = load_inventory(set_context["id"])
        if inventory_cards:
            return inventory_cards

    return load_cards_from_api(set_context)


def variation_label(card):
    label = f'{card["number"]} - {card["name"]}'

    finish = card["finish"].strip()

    if finish and finish.lower() != "normal":
        label += f" ({finish})"

    return label


# ==========================================================
# HELPERS
# ==========================================================


def blank_row(header):
    return [""] * len(header)


def set_field(row, cols, key, value):
    if value is None:
        return
    idx = cols.get(normalize_column_name(key))
    if idx is not None:
        row[idx] = value


def get_parent_custom_label(cards):
    if not cards:
        return ""
    sku = cards[0]["sku"].strip()
    return sku.split("-")[0].lower() if sku else ""


def build_output_path(set_context):

    return str(Path(OUTPUT_DIR) / f"{set_context['id'].upper()}_Ebay_Variation_Upload.csv")


def build_image_base_url(set_context):

    return (
        f"https://raw.githubusercontent.com/"
        f"{GITHUB_OWNER}/{GITHUB_REPO}/{GITHUB_BRANCH}/"
        f"cards/{set_context['id'].upper()}"
    )


def build_listing_title(set_context):

    return f"Pokemon TCG {set_context['name']} Bulk Singles - C, RH, HR"


def build_github_image_url(set_context, filename):

    filename = str(filename or "").strip()

    if not filename:
        return ""

    if filename.startswith("http"):
        return filename

    image_base_url = build_image_base_url(set_context)

    return image_base_url.rstrip("/") + "/" + filename


def build_cover_image_url(cards, set_context):

    if not cards:
        return ""

    for card in cards:
        image_url = build_github_image_url(set_context, card.get("image"))

        if image_url:
            return image_url

    return ""


# ==========================================================
# PARENT LISTING
# ==========================================================


def create_parent(cards, header, cols, set_context):
    parent = blank_row(header)
    cover_image_url = build_cover_image_url(cards, set_context)

    set_field(parent, cols, "Action(SiteID=AU|Country=AU|Currency=AUD|Version=1193|CC=UTF-8)", "Add")
    set_field(parent, cols, "Relationship", "")
    set_field(parent, cols, "Relationship details", "Card=" + ";".join(variation_label(card) for card in cards))
    set_field(parent, cols, "CustomLabel", get_parent_custom_label(cards))
    set_field(parent, cols, "Category", CATEGORY)
    set_field(parent, cols, "Title", build_listing_title(set_context))
    set_field(parent, cols, "ConditionID", CONDITION_ID)
    set_field(parent, cols, "CD:Card Condition - (ID: 40001)", CARD_CONDITION)
    set_field(parent, cols, "C:Card Condition", CARD_CONDITION_LABEL)
    set_field(parent, cols, "C:Game", PARENT_GAME)
    set_field(parent, cols, "PicURL", cover_image_url)
    set_field(parent, cols, "Description", DESCRIPTION)
    set_field(parent, cols, "Format", "FixedPrice")
    set_field(parent, cols, "Duration", "GTC")
    set_field(parent, cols, "ShippingProfileName", SHIPPING_PROFILE)
    set_field(parent, cols, "PaymentProfileName", PAYMENT_PROFILE)
    set_field(parent, cols, "ReturnProfileName", RETURN_PROFILE)
    set_field(parent, cols, "Location", LOCATION)
    set_field(parent, cols, "C:Set", set_context["name"])
    set_field(parent, cols, "C:Card Type", PARENT_CARD_TYPE)
    set_field(parent, cols, "C:Manufacturer", PARENT_MANUFACTURER)
    set_field(parent, cols, "C:Graded", PARENT_GRADED)
    set_field(parent, cols, "C:Card Size", PARENT_CARD_SIZE)
    set_field(parent, cols, "C:Language", PARENT_LANGUAGE)
    set_field(parent, cols, "C:Country/Region of Manufacture", PARENT_COUNTRY_MANUFACTURE)
    set_field(parent, cols, "C:Country of Origin", PARENT_COUNTRY_ORIGIN)
    set_field(parent, cols, "C:Age Level", PARENT_AGE_LEVEL)
    set_field(parent, cols, "C:Material", PARENT_MATERIAL)
    set_field(parent, cols, "PostalCode", PARENT_POSTAL_CODE)
    set_field(parent, cols, "WeightMajor", PARENT_WEIGHT_MAJOR)
    set_field(parent, cols, "WeightMinor", PARENT_WEIGHT_MINOR)
    set_field(parent, cols, "WeightUnit", PARENT_WEIGHT_UNIT)

    return parent


# ==========================================================
# CHILD LISTING
# ==========================================================


def create_child(card, header, cols, set_context):
    child = blank_row(header)

    set_field(child, cols, "Action(SiteID=AU|Country=AU|Currency=AUD|Version=1193|CC=UTF-8)", "Add")
    set_field(child, cols, "Relationship", "Variation")
    set_field(child, cols, "Relationship details", f"Card={variation_label(card)};")
    set_field(child, cols, "CustomLabel", card["sku"])
    set_field(child, cols, "StartPrice", format_price(card["price"]))
    set_field(child, cols, "Quantity", card["qty"])

    if card["image"]:
        image_url = build_github_image_url(set_context, card["image"])
        pic_value = f"{variation_label(card)}={image_url}"
        set_field(child, cols, "PicURL", pic_value)

    return child


# ==========================================================
# BUILD ALL CHILDREN
# ==========================================================


def create_children(cards, header, cols, set_context):
    return [create_child(card, header, cols, set_context) for card in cards]


def export_cards(cards, set_context, output_path=None, title_override=None):
    header, cols = load_template()
    cards = list(cards or [])
    total_cards = len(cards)
    processed_cards = 0
    output_path = output_path or build_output_path(set_context)

    grouped_cards = {}
    for card in cards:
        listing_group = str(card.get("listing_group") or "default")
        grouped_cards.setdefault(listing_group, []).append(card)

    rows = []
    for _, group_cards in grouped_cards.items():
        if not group_cards:
            continue

        parent = create_parent(group_cards, header, cols, set_context)
        group_title = title_override or group_cards[0].get("listing_title_override")
        if group_title:
            set_field(parent, cols, "Title", str(group_title))

        rows.append(parent)
        for card in group_cards:
            processed_cards += 1
            emit_csv_progress(processed_cards, total_cards, variation_label(card))
            rows.append(create_child(card, header, cols, set_context))

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with open(output_file, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)

    return {
        "output_file": str(output_file),
        "groups": len(grouped_cards),
        "cards": len(cards),
    }


# ==========================================================
# EXPORT
# ==========================================================


def export_csv(selected_set):
    set_context = get_set_config_by_id(selected_set)
    cards = load_cards_for_export(set_context)
    output_path = build_output_path(set_context)

    print(f"Found {len(cards)} cards")
    print("Building variations...")
    result = export_cards(cards, set_context, output_path=output_path)
    print(f"Created {result.get('cards', 0)} variations")

    print()
    print("====================================")
    print("Export Complete!")
    print("====================================")
    print(f"Cards Exported : {result.get('cards', 0)}")
    print(f"Output File    : {output_path}")
    print()


# ==========================================================
# MAIN
# ==========================================================


def main():
    try:
        if len(sys.argv) < 2:
            print("Usage:")
            print("python ebay_variation_exporter.py <SET_ID>")
            return

        export_csv(sys.argv[1])

    except FileNotFoundError as e:
        print()
        print("Missing file:")
        print(e)

    except PermissionError:
        print()
        print("Please close the CSV file before exporting.")

    except Exception as e:
        print()
        print("Unexpected Error")
        print(e)


if __name__ == "__main__":
    main()
